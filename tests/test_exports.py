import csv
from datetime import timedelta
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.db.session import SessionLocal
from app.models.client import Client
from app.models.common import utc_now
from app.models.crawl import CrawlRun, UrlLink, UrlSnapshot
from app.models.discovery import CrawlJob, Url, UrlSource
from app.models.exports import Export
from app.models.issues import Change, Issue
from app.models.jobs import JobListing
from app.models.recommendations import RecommendationTask
from app.models.user import User
from app.models.website import Website, WebsiteSettings
from app.services import exports as export_service


@pytest.mark.parametrize(
    "export_type,suffix",
    [("urls", "csv"), ("vacancies", "csv"), ("tasks", "csv"), ("excel", "xlsx")],
)
def test_generates_export(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    export_type: str,
    suffix: str,
) -> None:
    monkeypatch.setattr(export_service, "EXPORT_ROOT", tmp_path)
    with SessionLocal() as db:
        client = Client(name="Export client")
        website = Website(client=client, name="Export site", base_url="https://example.com")
        website.settings = WebsiteSettings()
        db.add(website)
        db.flush()
        db.add(Url(website_id=website.id, normalized_url="https://example.com/"))
        export = Export(website_id=website.id, export_type=export_type)
        db.add(export)
        db.commit()
        export_id = export.id

    export_service.generate_export(str(export_id))

    with SessionLocal() as db:
        completed = db.get(Export, export_id)
        assert completed and completed.status == "succeeded"
        path = Path(completed.file_path or "")
        assert path.suffix == f".{suffix}"
        assert path.stat().st_size > 0
        if export_type == "excel":
            assert "Vacancies" in load_workbook(path, read_only=True).sheetnames


def test_datasets_include_human_readable_urls() -> None:
    with SessionLocal() as db:
        client = Client(name="Readable export client")
        website = Website(client=client, name="Readable site", base_url="https://example.com")
        website.settings = WebsiteSettings()
        db.add(website)
        db.flush()
        source = Url(website_id=website.id, normalized_url="https://example.com/source")
        target = Url(website_id=website.id, normalized_url="https://example.com/target")
        db.add_all([source, target])
        db.flush()
        job = CrawlJob(website_id=website.id, job_type="full_site_crawl")
        db.add(job)
        db.flush()
        run = CrawlRun(
            crawl_job_id=job.id,
            website_id=website.id,
            crawl_type="full_site_crawl",
            status="succeeded",
            finished_at=utc_now(),
        )
        db.add(run)
        db.flush()
        old_job = CrawlJob(website_id=website.id, job_type="full_site_crawl")
        db.add(old_job)
        db.flush()
        old_run = CrawlRun(
            crawl_job_id=old_job.id,
            website_id=website.id,
            crawl_type="full_site_crawl",
            status="succeeded",
            finished_at=utc_now() - timedelta(days=1),
        )
        db.add(old_run)
        db.flush()
        snapshot = UrlSnapshot(
            url_id=target.id,
            crawl_run_id=run.id,
            requested_url=target.normalized_url,
        )
        db.add(snapshot)
        db.flush()
        db.add_all(
            [
                Issue(
                    website_id=website.id,
                    url_id=target.id,
                    issue_type="missing_title",
                    category="onpage",
                    severity="medium",
                    title="Title ontbreekt",
                    description="Test",
                    recommended_action="Herstel",
                ),
                Issue(
                    website_id=website.id,
                    url_id=target.id,
                    issue_type="job_posting_schema_missing",
                    category="structured_data",
                    severity="high",
                    title="Vacature mist JobPosting-schema",
                    description="Test",
                    recommended_action="Voeg schema toe",
                ),
                JobListing(
                    website_id=website.id,
                    url_id=target.id,
                    latest_snapshot_id=snapshot.id,
                    detection_sources=["url_pattern", "page_text"],
                    title="SEO specialist",
                    employer="Example",
                    locations=["Amsterdam"],
                    employment_types=["FULL_TIME"],
                    lifecycle_status="active",
                    current_status_code=200,
                    is_indexable=True,
                    inbound_internal_links=2,
                ),
                Change(
                    website_id=website.id,
                    url_id=target.id,
                    current_snapshot_id=snapshot.id,
                    change_type="new_url",
                ),
                UrlLink(
                    crawl_run_id=run.id,
                    source_url_id=source.id,
                    target_url=target.normalized_url,
                    target_url_id=target.id,
                    is_internal=True,
                    is_nofollow=False,
                ),
                UrlLink(
                    crawl_run_id=run.id,
                    source_url_id=source.id,
                    target_url=target.normalized_url,
                    target_url_id=target.id,
                    is_internal=True,
                    is_nofollow=False,
                ),
                UrlLink(
                    crawl_run_id=old_run.id,
                    source_url_id=source.id,
                    target_url=target.normalized_url,
                    target_url_id=target.id,
                    is_internal=True,
                    is_nofollow=False,
                ),
            ]
        )
        db.commit()

        datasets = export_service._datasets(db, website.id)

        assert datasets["urls"][0][0] == "url"
        assert datasets["issues"][0][0] == "url"
        assert datasets["issues"][1][0][0] == target.normalized_url
        assert datasets["changes"][0][0] == "url"
        assert datasets["changes"][1][0][0] == target.normalized_url
        assert datasets["links"][0][:2] == ["source_url", "target_url"]
        assert datasets["links"][1][0][0] == source.normalized_url
        assert datasets["vacancies"][0][0] == "url"
        assert datasets["vacancies"][1][0][0] == target.normalized_url
        assert datasets["vacancies"][1][0][1] == "SEO specialist"
        assert datasets["vacancies"][1][0][5] == "error"
        assert datasets["vacancies"][1][0][15] == "Vacature mist JobPosting-schema"
        assert all(
            not header.endswith("_id") for headers, _ in datasets.values() for header in headers
        )
        assert len(datasets["links"][1]) == 1


def test_filtered_url_export_contains_exact_selection_and_context(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    monkeypatch.setattr(export_service, "EXPORT_ROOT", tmp_path)
    with SessionLocal() as db:
        client = Client(name="Filtered export client")
        website = Website(client=client, name="Filtered site", base_url="https://example.com")
        website.settings = WebsiteSettings()
        db.add(website)
        db.flush()
        included = Url(website_id=website.id, normalized_url="https://example.com/included")
        excluded = Url(website_id=website.id, normalized_url="https://example.com/excluded")
        db.add_all([included, excluded])
        db.flush()
        export = Export(
            website_id=website.id,
            export_type="urls",
            item_ids=[str(included.id)],
            filters={"status": "2xx", "zoekopdracht": "included"},
        )
        db.add(export)
        db.commit()
        export_id = export.id

    export_service.generate_export(str(export_id))

    with SessionLocal() as db:
        completed = db.get(Export, export_id)
        assert completed and completed.file_path
        with Path(completed.file_path).open(encoding="utf-8", newline="") as handle:
            rows = list(csv.DictReader(handle))
        assert [row["url"] for row in rows] == ["https://example.com/included"]
        assert rows[0]["website"] == "Filtered site"
        assert '"status": "2xx"' in rows[0]["filters"]


def test_empty_filtered_selection_does_not_fall_back_to_all_rows() -> None:
    with SessionLocal() as db:
        client = Client(name="Empty selection client")
        website = Website(client=client, name="Empty site", base_url="https://example.com")
        website.settings = WebsiteSettings()
        db.add(website)
        db.flush()
        db.add(Url(website_id=website.id, normalized_url="https://example.com/unselected"))
        db.commit()

        datasets = export_service._datasets(db, website.id, selected_type="urls", item_ids=[])
        assert datasets["urls"][1] == []


def test_task_export_contains_assignment_and_exact_selection(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    monkeypatch.setattr(export_service, "EXPORT_ROOT", tmp_path)
    with SessionLocal() as db:
        client = Client(name="Task export client")
        website = Website(client=client, name="Task export site", base_url="https://example.com")
        website.settings = WebsiteSettings()
        user = User(
            email="editor@example.com",
            password_hash="test-password-hash",
            display_name="Content Editor",
            role="user",
        )
        db.add_all([website, user])
        db.flush()
        page = Url(website_id=website.id, normalized_url="https://example.com/page")
        db.add(page)
        db.flush()
        issue = Issue(
            website_id=website.id,
            url_id=page.id,
            issue_type="missing_title",
            category="onpage",
            severity="medium",
            title="Title ontbreekt",
            description="Test",
            recommended_action="Herstel",
        )
        db.add(issue)
        db.flush()
        included = RecommendationTask(
            website_id=website.id,
            assigned_to_user_id=user.id,
            primary_issue_id=issue.id,
            recommendation_type="write_title",
            definition_version="1",
            title="Schrijf paginatitel",
            category="onpage",
            primary_role="content_editor",
            priority="normal",
            priority_reason="Ontbrekende metadata",
            feasibility="straightforward",
            action="Schrijf een unieke titel.",
            rationale="De title ontbreekt.",
            acceptance_criteria=["Titel staat live"],
        )
        excluded = RecommendationTask(
            website_id=website.id,
            recommendation_type="other",
            definition_version="1",
            title="Andere taak",
            category="onpage",
            primary_role="seo_specialist",
            priority="low",
            priority_reason="Later",
            feasibility="straightforward",
            action="Controleer later.",
            rationale="Lagere prioriteit.",
        )
        db.add_all([included, excluded])
        db.flush()
        export = Export(
            website_id=website.id,
            export_type="tasks",
            item_ids=[str(included.id)],
            filters={"vakgebied": "Contentredactie"},
        )
        db.add(export)
        db.commit()
        export_id = export.id

    export_service.generate_export(str(export_id))

    with SessionLocal() as db:
        completed = db.get(Export, export_id)
        assert completed and completed.file_path
        with Path(completed.file_path).open(encoding="utf-8", newline="") as handle:
            rows = list(csv.DictReader(handle))
    assert [row["title"] for row in rows] == ["Schrijf paginatitel"]
    assert rows[0]["assigned_to"] == "Content Editor"
    assert rows[0]["issue_url"] == "https://example.com/page"
    assert rows[0]["acceptance_criteria"] == "Titel staat live"
    assert '"vakgebied": "Contentredactie"' in rows[0]["filters"]


def test_url_export_includes_current_and_historical_sources() -> None:
    with SessionLocal() as db:
        client = Client(name="Source export client")
        website = Website(client=client, name="Source site", base_url="https://example.com")
        website.settings = WebsiteSettings()
        db.add(website)
        db.flush()
        page = Url(website_id=website.id, normalized_url="https://example.com/page", crawl_depth=2)
        job = CrawlJob(website_id=website.id, job_type="full_site_crawl")
        db.add_all([page, job])
        db.flush()
        run = CrawlRun(
            crawl_job_id=job.id,
            website_id=website.id,
            crawl_type="full_site_crawl",
            status="succeeded",
            started_at=utc_now(),
        )
        db.add(run)
        db.flush()
        db.add_all(
            [
                UrlSource(
                    url_id=page.id,
                    source_type="internal_link",
                    source_url="https://example.com/",
                    last_seen_at=run.started_at,
                ),
                UrlSource(
                    url_id=page.id,
                    source_type="sitemap",
                    source_url="https://example.com/sitemap.xml",
                    last_seen_at=run.started_at - timedelta(days=1),
                ),
            ]
        )
        db.commit()
        headers, rows = export_service._datasets(db, website.id)["urls"]

    row = dict(zip(headers, rows[0], strict=True))
    assert row["crawl_depth"] == 2
    assert row["current_sources"] == "internal_link"
    assert row["historical_sources"] == "sitemap"
